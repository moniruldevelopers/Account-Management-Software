import pandas as pd
from django.contrib.auth.models import User
from django.shortcuts import render, redirect
from django.contrib.auth.hashers import make_password
from .forms import *
from .models import *
from django.shortcuts import render, get_object_or_404, redirect
from django.core.paginator import Paginator
from django.db.models import Q
from datetime import datetime, timedelta
from django.db.models import Sum,Count
from datetime import date
from django.utils.timezone import now
from .models import Transaction
from datetime import datetime
from django.contrib import messages
from django.utils.timezone import localtime
import pytz
from urllib.parse import quote
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
from io import BytesIO
from django.http import HttpResponse
from datetime import datetime
from django.contrib.auth.decorators import login_required
from django.contrib.auth.decorators import user_passes_test
from django.contrib.admin.models import LogEntry
from django.utils.safestring import mark_safe


@user_passes_test(lambda u: u.is_superuser or u.is_staff, login_url='login')
def upload_users(request):
    if request.method == 'POST':
        form = UploadExcelForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            new_users_count = 0
            existing_users_count = 0

            try:
                # Explicitly specify the openpyxl engine
                df = pd.read_excel(file, engine='openpyxl')

                for index, row in df.iterrows():
                    username = str(row['id']).strip()
                    full_name = row['full name'].strip()
                    phone_number = str(row['phone number']).strip() if 'phone number' in row and pd.notna(row['phone number']) else None
                    password = username  # Default password = username

                    # Create or get the User
                    user, created = User.objects.get_or_create(
                        username=username,
                        defaults={'first_name': full_name}
                    )

                    if created:
                        user.set_password(password)
                        user.save()
                        new_users_count += 1
                    else:
                        existing_users_count += 1

                    # Manually create the Profile if it doesn't exist
                    profile, _ = Profile.objects.get_or_create(user=user)

                    # Update the phone number if available
                    if phone_number:
                        profile.phone_number = phone_number
                        profile.save()

                messages.success(request, f'Upload completed! {new_users_count} new users added, {existing_users_count} users already existed.')
            except Exception as e:
                messages.error(request, f'Error processing file: {e}')

            return redirect('upload_users')
    else:
        form = UploadExcelForm()

    return render(request, 'user/upload_users.html', {'form': form})


@user_passes_test(lambda u: u.is_superuser or u.is_staff, login_url='login')
def download_upload_format(request):
    # Prepare data for the format (you can just create an empty or template row to download)
    data = {
        'id': [''],
        'full name': [''],
        'phone number': [''],
        'email': [''],
    }

    # Create a DataFrame with the template
    df = pd.DataFrame(data)

    # Generate Excel response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=upload_format.xlsx'

    # Save the DataFrame to Excel in memory
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Users')

    return response




@user_passes_test(lambda u: u.is_superuser or u.is_staff, login_url='login')
def home(request):
    # Get Bangladesh timezone
    bd_timezone = pytz.timezone("Asia/Dhaka")
    
    # Get the selected date from the request (or use today's date as default)
    filter_date_str = request.GET.get('filter_date')
    
    # If a filter date is provided, convert it to a datetime object
    if filter_date_str:
        filter_date = datetime.strptime(filter_date_str, "%Y-%m-%d").date()
    else:
        # Default to today's date
        filter_date = localtime().date()

    # Format the selected date for displaying in the template
    formatted_date = filter_date.strftime("%d %B, %Y")
    
    # Filter transactions by the selected date
    transactions = Transaction.objects.filter(created_date__date=filter_date)

    # Calculate totals for income, borrow, expense, and given
    total_income = sum(t.price for t in transactions if t.transaction_type == 'income')
    total_borrow = sum(t.price for t in transactions if t.transaction_type == 'borrow')
    total_expense = sum(t.price for t in transactions if t.transaction_type == 'expense')
    total_given = sum(t.price for t in transactions if t.transaction_type == 'given')

    # Calculate the available balance
    available_balance = (total_income + total_borrow) - (total_expense + total_given)

    # Pass the required context to the template
    context = {
        'transactions': transactions,
        'selected_date': filter_date.strftime("%Y-%m-%d"),  # For the input field in the filter
        'formatted_date': formatted_date,  # For displaying "14 March, 2025"
        'total_income': total_income,
        'total_borrow': total_borrow,
        'total_expense': total_expense,
        'total_given': total_given,
        'available_balance': available_balance,
    }
    return render(request, 'home.html', context)


@user_passes_test(lambda u: u.is_superuser or u.is_staff, login_url='login')
def add_edit_transaction(request, transaction_id=None):
    transaction = None

    if transaction_id:
        transaction = get_object_or_404(Transaction, id=transaction_id)

    if request.method == "POST":
        form = TransactionForm(request.POST, instance=transaction)
        if form.is_valid():
            saved_transaction = form.save(commit=False)

            # Assign created_by only if it's a new transaction
            if not saved_transaction.created_by:
                saved_transaction.created_by = request.user  

            saved_transaction.save()

            user_full_name = saved_transaction.transaction_by.get_full_name() or saved_transaction.transaction_by.username
            messages.success(request, f"Transaction added successfully for {user_full_name}!")

            # Redirect to the print view
            return redirect('transaction_print', transaction_id=saved_transaction.id)

    else:
        form = TransactionForm(instance=transaction)

    form.fields['transaction_by'].queryset = User.objects.all()
    form.fields['transaction_by'].label_from_instance = lambda obj: f"{obj.username} - {obj.get_full_name() or obj.username}"

    return render(request, 'transactions/add_transaction.html', {'form': form, 'transaction': transaction})

@user_passes_test(lambda u: u.is_superuser or u.is_staff, login_url='login')
def transaction_print(request, transaction_id):
    transaction = get_object_or_404(Transaction, id=transaction_id)
    return render(request, 'transactions/transaction_print.html', {'transaction': transaction})

@user_passes_test(lambda u: u.is_superuser or u.is_staff, login_url='login')
def transaction_list(request):
    search_query = request.GET.get('search', '')

    # Create a base queryset, ordering by latest first
    transactions = Transaction.objects.all().order_by('-id')  # Sorting by latest first

    # Apply search filter if search query is present
    if search_query:
        transactions = transactions.filter(
            Q(transaction_id__icontains=search_query) |  # Search by transaction ID
            Q(transaction_by__username__icontains=search_query) |
            Q(transaction_by__first_name__icontains=search_query) |
            Q(transaction_by__last_name__icontains=search_query) |
            Q(price__icontains=search_query) |
            Q(invoice_id__icontains=search_query) |
            Q(check_no__icontains=search_query) |
            Q(transaction_by__profile__phone_number__icontains=search_query)  # Search by phone number
        )

        if transactions.exists():
            messages.success(request, f"Found {transactions.count()} matching transaction(s).")
        else:
            messages.warning(request, "No matching transactions found.")

    # Pagination: Show 10 transactions per page
    paginator = Paginator(transactions, 10)  # Show 10 transactions per page
    page_number = request.GET.get('page')
    transactions = paginator.get_page(page_number)

    context = {
        'transactions': transactions,
        'search_query': search_query,
    }

    return render(request, 'transactions/transaction_list.html', context)

@user_passes_test(lambda u: u.is_superuser or u.is_staff, login_url='login')
def delete_transaction(request, transaction_id):
    transaction = get_object_or_404(Transaction, id=transaction_id)
    
    # Get the full name of the user who created the transaction
    user_full_name = transaction.transaction_by.get_full_name() or transaction.transaction_by.username

    # Delete the transaction
    transaction.delete()

    # Add the success message with the user's full name
    messages.success(request, f"Transaction deleted successfully for {user_full_name}.")

    # Redirect to the transaction list page
    return redirect('transaction_list')

@user_passes_test(lambda u: u.is_superuser or u.is_staff, login_url='login')
def user_list(request):
    query = request.GET.get('q', '')

    users = User.objects.all().order_by('-date_joined') # Order by date_joined descending
    if query:
        users = users.filter(username__icontains=query)

    admin_users = users.filter(is_staff=True)
    regular_users = users.filter(is_staff=False, is_superuser=False)

    # Pagination for all users
    paginator = Paginator(users, 10)
    page_number = request.GET.get('page')
    users_page = paginator.get_page(page_number)

    # Separate pagination for regular users
    regular_paginator = Paginator(regular_users, 10)
    regular_page_number = request.GET.get('regular_page')
    regular_users_page = regular_paginator.get_page(regular_page_number)

    context = {
        'users': users_page,
        'admin_users': admin_users,
        'regular_users': regular_users_page,
        'total_users': users.count(),
        'total_admin_users': admin_users.count(),
        'total_regular_users': regular_users.count(),
        'query': query,
    }

    return render(request, 'user/user_list.html', context)

@user_passes_test(lambda u: u.is_superuser or u.is_staff, login_url='login')
def edit_user(request, user_id):
    user = get_object_or_404(User, id=user_id)
    profile, created = Profile.objects.get_or_create(user=user)

    if request.method == 'POST':
        form = ProfileForm(request.POST, request.FILES, instance=profile) #add request.FILES
        if form.is_valid():
            form.save()
            user.first_name = request.POST.get('first_name', user.first_name)
            user.email = request.POST.get('email', user.email)
            user.save()
            messages.success(request, 'User updated successfully!')
            return redirect('user_list')
    else:
        form = ProfileForm(instance=profile)

    return render(request, 'user/edit_user.html', {'user': user, 'form': form})


@user_passes_test(lambda u: u.is_superuser or u.is_staff, login_url='login')
def delete_user(request, user_id):
    user = get_object_or_404(User, id=user_id)
    user.delete()
    messages.success(request, 'User deleted successfully!')
    return redirect('user_list')

@user_passes_test(lambda u: u.is_superuser or u.is_staff, login_url='login')
def reset_password(request, user_id):
    user = get_object_or_404(User, id=user_id)
    new_password = user.username  # Reset password to default (username)
    user.password = make_password(new_password)
    user.save()
    messages.success(request, f'Password reset successfully! New password: {new_password}')
    return redirect('user_list')

@user_passes_test(lambda u: u.is_superuser or u.is_staff, login_url='login')
def yearly_balance(request):
    # Get available years dynamically from the transactions
    available_years = Transaction.objects.dates('created_date', 'year', order='DESC').distinct()
    available_years = [year.year for year in available_years]

    # Get the selected year from GET request or default to the most recent available year
    selected_year = request.GET.get('year', available_years[0] if available_years else datetime.now().year)
    selected_year = int(selected_year)

    # Initialize monthly data storage
    monthly_balance = []
    total_yearly_balance = 0

    # Loop through each month (1 to 12)
    for month in range(1, 13):
        # Get total income for the month
        total_income = Transaction.objects.filter(
            transaction_type='income',
            created_date__year=selected_year,
            created_date__month=month
        ).aggregate(Sum('price'))['price__sum'] or 0

        # Get total borrow for the month
        total_borrow = Transaction.objects.filter(
            transaction_type='borrow',
            created_date__year=selected_year,
            created_date__month=month
        ).aggregate(Sum('price'))['price__sum'] or 0

        # Get total expense for the month
        total_expense = Transaction.objects.filter(
            transaction_type='expense',
            created_date__year=selected_year,
            created_date__month=month
        ).aggregate(Sum('price'))['price__sum'] or 0

        # Get total given for the month
        total_given = Transaction.objects.filter(
            transaction_type='given',
            created_date__year=selected_year,
            created_date__month=month
        ).aggregate(Sum('price'))['price__sum'] or 0

        # Calculate available balance
        available_balance = (total_income + total_borrow) - (total_expense + total_given)

        # Add to yearly total
        total_yearly_balance += available_balance

        # Store data for the month
        monthly_balance.append({
            'month': datetime(selected_year, month, 1).strftime('%B'),
            'income': total_income,
            'borrow': total_borrow,
            'expense': total_expense,
            'given': total_given,
            'available_balance': available_balance
        })

    context = {
        'selected_year': selected_year,
        'monthly_balance': monthly_balance,
        'available_years': available_years,
        'total_yearly_balance': total_yearly_balance
    }

    return render(request, 'transactions/yearly_balance.html', context)

@user_passes_test(lambda u: u.is_superuser or u.is_staff, login_url='login')
def yearly_given_borrow(request):
    # Get available years dynamically
    available_years = Transaction.objects.filter(transaction_type__in=['given', 'borrow']).dates('created_date', 'year', order='DESC').distinct()
    available_years = [year.year for year in available_years]

    # Get selected year
    selected_year = request.GET.get('year', available_years[0] if available_years else datetime.now().year)
    selected_year = int(selected_year)

    # Initialize monthly transaction list
    monthly_transactions = []

    for month in range(1, 13):
        transactions = Transaction.objects.filter(
            transaction_type__in=['given', 'borrow'],
            created_date__year=selected_year,
            created_date__month=month
        ).select_related('transaction_by')  # Fetch related User objects

        monthly_transactions.append({
            'month': datetime(selected_year, month, 1).strftime('%B'),
            'transactions': transactions
        })

    context = {
        'selected_year': selected_year,
        'monthly_transactions': monthly_transactions,
        'available_years': available_years
    }

    return render(request, 'transactions/yearly_given_borrow.html', context)





@user_passes_test(lambda u: u.is_superuser or u.is_staff, login_url='login')
def category_summary(request):
    categories = TransactionCategory.objects.all()
    category_data = []

    # Get year and month from request, default to current year and month
    current_datetime = datetime.now()
    selected_year = request.GET.get('year', str(current_datetime.year))
    selected_month = request.GET.get('month', str(current_datetime.month))

    transactions = Transaction.objects.all()

    # Convert to integers for comparison
    try:
        selected_year = int(selected_year)
    except ValueError:
        selected_year = current_datetime.year

    try:
        selected_month = int(selected_month)
    except ValueError:
        selected_month = current_datetime.month

    # Filter transactions based on created_date's year and month
    transactions = transactions.filter(created_date__year=selected_year, created_date__month=selected_month)

    for category in categories:
        total_income = transactions.filter(category=category, transaction_type='income').aggregate(total=Sum('price'))['total'] or 0
        total_expense = transactions.filter(category=category, transaction_type='expense').aggregate(total=Sum('price'))['total'] or 0

        balance = total_income - total_expense

        category_data.append({
            'category': category.name,
            'total_income': total_income,
            'total_expense': total_expense,
            'balance': balance,
        })

    # Get distinct years from created_date
    years = Transaction.objects.values('created_date__year').distinct().order_by('-created_date__year')

    # Prepare months
    months = [{'id': i, 'name': datetime(1900, i, 1).strftime('%B')} for i in range(1, 13)]

    #Add success message
    month_name = ""
    for month in months:
        if month['id'] == selected_month:
            month_name = month['name']
            break;

    messages.success(request, f"Summary for {month_name}, {selected_year}.")

    return render(request, 'transactions/category_summary.html', {
        'category_data': category_data,
        'years': years,
        'months': months,
        'selected_year': selected_year,
        'selected_month': selected_month,
    })

@user_passes_test(lambda u: u.is_superuser or u.is_staff, login_url='login')
def category_monthly_transactions(request, category_name):
    # Get year and month from request, default to current year and month
    current_datetime = datetime.now()
    selected_year = request.GET.get('year', str(current_datetime.year))
    selected_month = request.GET.get('month', str(current_datetime.month))

    try:
        selected_year = int(selected_year)
    except ValueError:
        selected_year = current_datetime.year

    try:
        selected_month = int(selected_month)
    except ValueError:
        selected_month = current_datetime.month

    category = get_object_or_404(TransactionCategory, name=category_name)
    transactions = Transaction.objects.filter(
        category=category,
        created_date__year=selected_year,
        created_date__month=selected_month,
    )

    total_transactions = transactions.count()

    paginator = Paginator(transactions, 300)  # Show 10 transactions per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Export to Excel functionality
    if request.GET.get('export') == 'true':
        wb = Workbook()
        ws = wb.active
        ws.title = 'Transactions'

        headers = ['Transaction ID', 'Transaction Type', 'Full Name', 'Price', 'Description', 'Date']
        ws.append(headers)

        for cell in ws[1]:
            cell.font = Font(bold=True)

        for transaction in transactions:
            created_date = transaction.created_date
            if created_date.tzinfo is not None:
                created_date = created_date.replace(tzinfo=None)

            ws.append([
                transaction.transaction_id,
                transaction.transaction_type,
                f"{transaction.transaction_by.get_full_name()} ({transaction.transaction_by.username})",
                transaction.price,
                transaction.description,
                created_date,
            ])

        excel_month = datetime(1900, selected_month, 1).strftime('%B')
        filename = f"{excel_month}_{category_name}_{selected_year}.xlsx"

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={filename}'
        wb.save(response)
        return response

    return render(request, 'transactions/category_monthly_transactions.html', {
        'category': category_name,
        'page_obj': page_obj,
        'total_transactions': total_transactions,
        'selected_year': selected_year,
        'selected_month': selected_month,
    })



@user_passes_test(lambda u: u.is_superuser or u.is_staff, login_url='login')
def category_transactions_filter(request):
    categories = TransactionCategory.objects.all()  # Get all categories for selection
    selected_category_name = request.GET.get('category')
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')

    transactions = Transaction.objects.none()  # Default to empty queryset

    if selected_category_name and start_date_str and end_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            selected_category = get_object_or_404(TransactionCategory, name=selected_category_name)
            transactions = Transaction.objects.filter(
                category=selected_category,
                created_date__date__range=[start_date, end_date],
            )

            # Calculate the number of days between start and end date
            delta = end_date - start_date
            total_days = delta.days + 1  # Include both start and end dates

            messages.success(request, f"Transactions filtered from {start_date} to {end_date} ({total_days} days).")

        except (ValueError, TransactionCategory.DoesNotExist):
            # Handle invalid date formats or category not found
            messages.error(request, "Invalid date format or category not found.") #Added error message
            pass  # Keep transactions as empty queryset

    total_transactions = transactions.count()
    paginator = Paginator(transactions, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    if request.GET.get('export') == 'true' and transactions:
        wb = Workbook()
        ws = wb.active
        ws.title = 'Transactions'

        headers = ['Transaction ID', 'Transaction Type', 'Full Name', 'Price', 'Description', 'Date']
        ws.append(headers)

        for cell in ws[1]:
            cell.font = Font(bold=True)

        for transaction in transactions:
            created_date = transaction.created_date
            if created_date.tzinfo is not None:
                created_date = created_date.replace(tzinfo=None)

            ws.append([
                transaction.transaction_id,
                transaction.transaction_type,
                f"{transaction.transaction_by.get_full_name()} ({transaction.transaction_by.username})",
                transaction.price,
                transaction.description,
                created_date,
            ])

        filename = f"Transactions_{selected_category_name}_{start_date_str}_{end_date_str}.xlsx" if selected_category_name and start_date_str and end_date_str else "Transactions.xlsx"

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={filename}'
        wb.save(response)
        return response

    return render(request, 'transactions/category_transactions_filter.html', {
        'categories': categories,
        'selected_category_name': selected_category_name,
        'start_date_str': start_date_str,
        'end_date_str': end_date_str,
        'page_obj': page_obj,
        'total_transactions': total_transactions,
    })







@user_passes_test(lambda u: u.is_superuser or u.is_staff, login_url='login')
def category_list(request):
    categories = TransactionCategory.objects.all()
    return render(request, 'transactions/category_list.html', {'categories': categories})

@user_passes_test(lambda u: u.is_superuser or u.is_staff, login_url='login')
def add_or_edit_transaction_category(request, category_id=None):
    # If category_id is provided, it's an edit; otherwise, it's an add
    if category_id:
        category = get_object_or_404(TransactionCategory, id=category_id)
    else:
        category = None

    if request.method == 'POST':
        # Handle the form submission to add or update the category
        form = TransactionCategoryForm(request.POST, instance=category)
        if form.is_valid():
            form.save()
            return redirect('category_list')  # Redirect to the category list after saving
    else:
        # For GET requests, show the existing category data in the form or an empty form
        form = TransactionCategoryForm(instance=category)

    return render(request, 'transactions/add_or_edit_category.html', {'form': form, 'category': category})






@user_passes_test(lambda u: u.is_superuser or u.is_staff, login_url='login')
def transaction_filter(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Ensure that start_date and end_date are valid dates
    try:
        if start_date:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        if end_date:
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    except ValueError:
        start_date = None
        end_date = None

    # Apply filters
    transactions = Transaction.objects.all()

    if start_date:
        transactions = transactions.filter(created_date__date__gte=start_date)  # compare only the date part
    if end_date:
        transactions = transactions.filter(created_date__date__lte=end_date)  # compare only the date part

    # Pagination
    paginator = Paginator(transactions, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Calculate income, expense, and available balance
    income = transactions.filter(transaction_type__in=['income', 'borrow']).aggregate(total_income=Sum('price'))['total_income'] or 0
    expense = transactions.filter(transaction_type__in=['expense', 'given']).aggregate(total_expense=Sum('price'))['total_expense'] or 0
    available_balance = income - expense

    # Message to display which dates are filtered and transaction count
    date_message = ""
    if start_date and end_date:
        date_message = f"Showing results from {start_date} to {end_date}. Total transactions: {transactions.count()}."
    elif start_date:
        date_message = f"Showing results from {start_date}. Total transactions: {transactions.count()}."
    elif end_date:
        date_message = f"Showing results up to {end_date}. Total transactions: {transactions.count()}."
    else:
        date_message = f"Showing all transactions. Total transactions: {transactions.count()}."

    context = {
        'page_obj': page_obj,
        'start_date': start_date,
        'end_date': end_date,
        'income': income,
        'expense': expense,
        'available_balance': available_balance,
        'date_message': date_message,
    }

    return render(request, 'transactions/transaction_filter.html', context)


@user_passes_test(lambda u: u.is_superuser or u.is_staff, login_url='login')
def user_profile(request, username):
    user = get_object_or_404(User, username=username)
    profile = get_object_or_404(Profile, user=user)
    transactions = Transaction.objects.filter(transaction_by=user).order_by('-created_date')

    if request.GET.get('export') == 'true':
        return export_transactions_to_excel(user, transactions)

    paginator = Paginator(transactions, 300)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'transactions/user_profile.html', {
        'user': user,
        'profile': profile,
        'page_obj': page_obj,
    })
@user_passes_test(lambda u: u.is_superuser or u.is_staff, login_url='login')
def search_user(request):
    query = request.GET.get('q', '').strip()

    if not query:
        return render(request, 'transactions/user_not_found.html')

    try:
        user = User.objects.get(username=query)
    except User.DoesNotExist:
        return render(request, 'transactions/user_not_found.html')

    transactions = Transaction.objects.filter(transaction_by=user).order_by('-created_date')

    if request.GET.get('export') == 'true':
        return export_transactions_to_excel(user, transactions)

    return render(request, 'transactions/user_profile.html', {
        'user': user,
        'profile': user.profile,
        'page_obj': transactions,
    })

@user_passes_test(lambda u: u.is_superuser or u.is_staff, login_url='login')
def export_transactions_to_excel(user, transactions):
    wb = Workbook()
    ws = wb.active
    ws.title = f"{user.username}_Transactions"

    headers = ['Transaction ID', 'Type', 'Category', 'Price', 'Description', 'Date', 'Created By']
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    for transaction in transactions:
        created_date = transaction.created_date.strftime('%Y-%m-%d')

        ws.append([
            transaction.transaction_id,
            transaction.transaction_type,
            transaction.category.name if transaction.category else 'N/A',
            transaction.price,
            transaction.description,
            created_date,
            f"{transaction.created_by.username} ({transaction.created_by.get_full_name()})" if transaction.created_by else "System"
        ])

    full_name = user.get_full_name().replace(" ", "_") if user.get_full_name() else user.username
    filename = f"{full_name}_{user.username}_transactions.xlsx"

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    wb.save(response)
    return response



