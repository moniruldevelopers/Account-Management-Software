import openpyxl
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.contrib import messages
from .models import Book
from .forms import *
from django.contrib import messages
from django.core.paginator import Paginator
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy
from django.contrib.auth.decorators import user_passes_test


@user_passes_test(lambda u: u.is_superuser or u.is_staff, login_url='login')
def book_upload_view(request):
    if request.method == "POST":
        form = BookUploadForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES["file"]
            wb = openpyxl.load_workbook(file)
            sheet = wb.active

            uploaded_count = 0
            duplicate_count = 0

            for row in sheet.iter_rows(min_row=2, values_only=True):
                book_name, writer, publication, quantity = row  # Match model fields

                if book_name:  # Ensure book name is provided
                    obj, created = Book.objects.get_or_create(
                        name=book_name,
                        defaults={
                            "writer": writer,
                            "publication": publication,
                            "quantity": quantity or 1,  # Default quantity to 1 if empty
                        },
                    )

                    if created:
                        uploaded_count += 1
                    else:
                        duplicate_count += 1

            messages.success(request, f"Upload Complete: {uploaded_count} new books added, {duplicate_count} duplicates ignored.")
            return redirect("book_upload")

    else:
        form = BookUploadForm()

    return render(request, "library/book_upload.html", {"form": form})

@user_passes_test(lambda u: u.is_superuser or u.is_staff, login_url='login')
def download_book_format(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Books Format"

    # Define headers matching model fields
    headers = ["Book Name", "Writer", "Publication", "Quantity"]
    ws.append(headers)

    # Make headers bold
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="book_upload_format.xlsx"'

    wb.save(response)
    return response







@user_passes_test(lambda u: u.is_superuser or u.is_staff, login_url='login')
def book_list_view(request):
    books = Book.objects.all()
    total_books = books.count()  # Count total books
    total_quantity = sum(book.quantity for book in books)  # Sum of all book quantities

    # Pagination
    paginator = Paginator(books, 1)  # Show 10 books per page
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    return render(request, "library/book_list.html", {
        "page_obj": page_obj,
        "total_books": total_books,
        "total_quantity": total_quantity,
    })

@user_passes_test(lambda u: u.is_superuser or u.is_staff, login_url='login')
def export_books_to_excel(request):
    books = Book.objects.all()
    total_books = books.count()
    total_quantity = sum(book.quantity for book in books)

    # Create an Excel workbook and sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Books List"

    # Add headers
    headers = ["#", "Book Name", "Writer", "Publication", "Quantity"]
    ws.append(headers)

    # Populate data
    for index, book in enumerate(books, start=1):
        ws.append([index, book.name, book.writer or "Unknown", book.publication or "Unknown", book.quantity])

    # Add total books and total quantity at the bottom
    ws.append([])  # Empty row for separation
    ws.append(["Total Books", total_books])
    ws.append(["Total Quantity", total_quantity])

    # Prepare response
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="books_list.xlsx"'
    wb.save(response)
    
    return response






class LibraryRecordListView(ListView):
    model = LibraryRecord
    template_name = 'library/libraryrecord_list.html'
    context_object_name = 'records'
    paginate_by = 10
    ordering = ['-created_date']  # Order by created_date descending (latest first)

    def get_queryset(self):
        # You can customize the queryset here if needed. Make sure to use the ordering.
        return LibraryRecord.objects.all().order_by('-created_date')

class LibraryRecordCreateView(CreateView):
    model = LibraryRecord
    form_class = LibraryRecordForm
    template_name = 'library/add_library_record_form.html'
    success_url = reverse_lazy('libraryrecord_list')

    def form_valid(self, form):
        # Set created_by to the current logged-in user
        form.instance.created_by = self.request.user  # Assign the User instance, not a string
        
        # Continue saving the form as usual
        return super().form_valid(form)

class LibraryRecordUpdateView(UpdateView):
    model = LibraryRecord
    form_class = LibraryRecordForm
    template_name = 'library/add_library_record_form.html'
    success_url = reverse_lazy('libraryrecord_list')

class LibraryRecordDeleteView(DeleteView):
    model = LibraryRecord
    template_name = 'library/libraryrecord_confirm_delete.html'
    success_url = reverse_lazy('libraryrecord_list')